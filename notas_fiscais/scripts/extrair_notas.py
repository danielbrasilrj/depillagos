#!/usr/bin/env python3
"""
Extrai notas fiscais com cota parte de profissional parceiro via API NFS-e Nacional.
Usa certificado digital A1 (.pfx) para autenticação mTLS.
Complementa com PDFs locais para notas ainda não distribuídas no ADN.

Uso:
    python3 scripts/extrair_notas.py
    python3 scripts/extrair_notas.py --periodo 2026-02
    python3 scripts/extrair_notas.py --output resultado.csv
    python3 scripts/extrair_notas.py --pdf "Notas (1).pdf" "Notas (2).pdf"  # complementar com PDFs
"""

import argparse
import base64
import glob
import gzip
import json
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from xml.etree import ElementTree as ET

# === CONFIGURAÇÃO ===
PROJETO_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PFX_PATH = os.path.join(PROJETO_DIR, "certificado", "DEPILAGOS MATRIZ 2025.pfx")
PFX_PASSWORD = None  # Definido via --senha ou variável DEPILAGOS_PFX_PASSWORD
CNPJ_DEPILAGOS = "09223558000100"
API_BASE = "https://adn.nfse.gov.br/contribuintes"
NS = {"nfse": "http://www.sped.fazenda.gov.br/nfse"}
MESES_ABREV = {
    "01": "jan", "02": "fev", "03": "mar", "04": "abr",
    "05": "mai", "06": "jun", "07": "jul", "08": "ago",
    "09": "set", "10": "out", "11": "nov", "12": "dez",
}


def extrair_cert_key(pfx_path, password):
    """Extrai certificado e chave privada do arquivo .pfx para arquivos temporários."""
    cert_path = os.path.join(tempfile.gettempdir(), "depilagos_cert.pem")
    key_path = os.path.join(tempfile.gettempdir(), "depilagos_key.pem")

    subprocess.run(
        ["openssl", "pkcs12", "-in", pfx_path, "-clcerts", "-nokeys",
         "-out", cert_path, "-passin", f"pass:{password}", "-legacy"],
        capture_output=True, check=True,
    )
    subprocess.run(
        ["openssl", "pkcs12", "-in", pfx_path, "-nocerts", "-nodes",
         "-out", key_path, "-passin", f"pass:{password}", "-legacy"],
        capture_output=True, check=True,
    )
    return cert_path, key_path


def consultar_lote(cert_path, key_path, nsu_inicio):
    """Consulta um lote de documentos a partir do NSU informado."""
    url = f"{API_BASE}/DFe/{nsu_inicio}?cnpjConsulta={CNPJ_DEPILAGOS}&lote=true"
    result = subprocess.run(
        ["curl", "-sS", "--max-time", "60",
         "--cert", cert_path, "--key", key_path,
         "-H", "Accept: application/json", url],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        print(f"Erro curl: {result.stderr}", file=sys.stderr)
        return None
    return json.loads(result.stdout)


def decodificar_xml(arquivo_xml_b64):
    """Decodifica o XML compactado (GZip + Base64) retornado pela API."""
    xml_bytes = gzip.decompress(base64.b64decode(arquivo_xml_b64))
    return xml_bytes.decode("utf-8")


def extrair_dados_nota(doc):
    """Extrai dados relevantes de um documento do lote da API."""
    xml_str = decodificar_xml(doc["ArquivoXml"])
    root = ET.fromstring(xml_str)

    # Número da NFS-e
    nnfse_el = root.find(".//nfse:nNFSe", NS)
    nnfse = nnfse_el.text if nnfse_el is not None else ""

    # Data de competência (data do serviço)
    dcompet_el = root.find(".//nfse:dCompet", NS)
    dcompet = dcompet_el.text if dcompet_el is not None else ""

    # Data de emissão
    dhemi_el = root.find(".//nfse:dhEmi", NS)
    dhemi = dhemi_el.text[:10] if dhemi_el is not None else ""

    # Data de processamento
    dhproc_el = root.find(".//nfse:dhProc", NS)
    dhproc = dhproc_el.text[:10] if dhproc_el is not None else ""

    # Valor do serviço
    vserv_el = root.find(".//nfse:vServ", NS)
    vserv = float(vserv_el.text) if vserv_el is not None else 0.0

    # Descrição do serviço
    desc_el = root.find(".//nfse:xDescServ", NS)
    desc = desc_el.text if desc_el is not None else ""

    # Chave de acesso
    chave = doc.get("ChaveAcesso", "")

    # Status da nota
    cstat_el = root.find(".//nfse:cStat", NS)
    cstat = cstat_el.text if cstat_el is not None else "100"

    # Extrair cotas do rateio - dois formatos possíveis
    # parceiros: cnpj -> {"cota_parceiro": valor, "cota_salao": valor}
    parceiros = {}

    if "COTA-PARTE" in desc:
        # Formato 1 (série 49999): SALAO-PARCEIRO: CNPJ COTA-PARTE R$valor
        # Neste formato, cada serviço tem uma linha SALAO-PARCEIRO e uma PROFISSIONAL-PARCEIRO
        # A cota do salão é global (mesmo CNPJ Depilagos), então dividimos proporcionalmente
        cota_salao_total = 0.0
        parceiros_raw = {}
        cotas = re.findall(
            r"(SALAO-PARCEIRO|PROFISSIONAL-PARCEIRO):\s*(\d{11,14})\s+COTA-PARTE\s+R\$([\d.]+(?:,\d+)?)",
            desc,
        )
        for tipo, cnpj, valor in cotas:
            valor_f = float(valor.replace(",", "."))
            if tipo == "SALAO-PARCEIRO":
                cota_salao_total += valor_f
            elif tipo == "PROFISSIONAL-PARCEIRO":
                parceiros_raw[cnpj] = parceiros_raw.get(cnpj, 0.0) + valor_f

        # Se só um parceiro, cota_salao = total. Se múltiplos, dividir proporcionalmente.
        total_parceiros = sum(parceiros_raw.values())
        for cnpj, cota_p in parceiros_raw.items():
            proporcao = cota_p / total_parceiros if total_parceiros > 0 else 1.0
            parceiros[cnpj] = {
                "cota_parceiro": cota_p,
                "cota_salao": cota_salao_total * proporcao,
            }

    elif "Rateio referente" in desc:
        # Formato 2 (série 1): CNPJ: DIGITS - NOME - R$ valor
        # Cada bloco "Rateio referente a" tem um par salão/parceiro
        blocos = re.split(r"Rateio referente a Salao/Profissional parceiro:", desc)
        for bloco in blocos[1:]:
            cnpjs = re.findall(
                r"CNPJ:\s*(\d{11,14})\b.{1,150}?R\$\s*([\d]+[.,][\d]+)",
                bloco,
            )
            cota_salao_bloco = 0.0
            cnpj_parceiro = ""
            cota_parceiro = 0.0

            for cnpj, valor in cnpjs:
                valor_f = float(valor.replace(".", "").replace(",", "."))
                if cnpj == CNPJ_DEPILAGOS:
                    cota_salao_bloco += valor_f
                else:
                    cnpj_parceiro = cnpj
                    cota_parceiro += valor_f

            if cnpj_parceiro:
                if cnpj_parceiro in parceiros:
                    parceiros[cnpj_parceiro]["cota_parceiro"] += cota_parceiro
                    parceiros[cnpj_parceiro]["cota_salao"] += cota_salao_bloco
                else:
                    parceiros[cnpj_parceiro] = {
                        "cota_parceiro": cota_parceiro,
                        "cota_salao": cota_salao_bloco,
                    }

    return {
        "nfse": nnfse,
        "chave_acesso": chave,
        "data_servico": dcompet,
        "data_emissao": dhproc,
        "valor_total": vserv,
        "descricao": desc,
        "parceiros": parceiros,
        "tem_rateio": len(parceiros) > 0,
        "cstat": cstat,
    }


def extrair_notas_pdf(pdf_paths):
    """Extrai notas com rateio de arquivos PDF (complemento para notas não disponíveis na API)."""
    notas = []
    for pdf_path in pdf_paths:
        print(f"  Processando PDF: {os.path.basename(pdf_path)}...")
        tmp_txt = os.path.join(tempfile.gettempdir(), "notas_pdf_tmp.txt")
        subprocess.run(["pdftotext", pdf_path, tmp_txt], capture_output=True, check=True)

        with open(tmp_txt, encoding="utf-8") as f:
            text = f.read()

        pages = text.split("Nota Fiscal de Serviço Eletrônica")
        pages = [p for p in pages if p.strip()]

        for page in pages:
            if "Rateio" not in page:
                continue

            nfse_match = re.search(r"Número da NFS-e:\s*(\d+)", page)
            nfse_num = nfse_match.group(1) if nfse_match else ""

            data_servico_match = re.search(
                r"(\d{2}/\d{2}/\d{4})\s*\n\s*\n?\s*\w{9,}", page
            )
            data_servico = data_servico_match.group(1) if data_servico_match else ""

            # Converter DD/MM/YYYY para YYYY-MM-DD para consistência interna
            data_servico_iso = ""
            if data_servico and len(data_servico) == 10:
                d, m, y = data_servico.split("/")
                data_servico_iso = f"{y}-{m}-{d}"

            data_emissao = "27/02/2026"  # default from PDF metadata

            full_text = " ".join(page.split())
            matches = re.findall(
                r"CNPJ:\s*(\d{11,14})\b.{1,150}?R\$\s*([\d]+[.,][\d]+)", full_text
            )

            cota_salao = 0.0
            parceiros = {}

            for cnpj, valor in matches:
                valor_f = float(valor.replace(".", "").replace(",", "."))
                if cnpj == CNPJ_DEPILAGOS:
                    cota_salao += valor_f
                else:
                    parceiros[cnpj] = parceiros.get(cnpj, 0.0) + valor_f

            if parceiros:
                # Converter para novo formato: cnpj -> {"cota_parceiro", "cota_salao"}
                parceiros_fmt = {}
                for cnpj, cota_p in parceiros.items():
                    proporcao = cota_p / sum(parceiros.values()) if sum(parceiros.values()) > 0 else 1.0
                    parceiros_fmt[cnpj] = {
                        "cota_parceiro": cota_p,
                        "cota_salao": cota_salao * proporcao,
                    }
                notas.append({
                    "nfse": nfse_num,
                    "chave_acesso": "",
                    "data_servico": data_servico_iso,
                    "data_emissao": data_emissao,
                    "valor_total": 0.0,
                    "descricao": "",
                    "parceiros": parceiros_fmt,
                    "tem_rateio": True,
                    "nsu": 0,
                    "fonte": "pdf",
                })

    print(f"  PDFs: {len(notas)} notas com rateio extraídas")
    return notas


def formatar_cnpj(cnpj):
    """Formata CNPJ de 14 dígitos para XX.XXX.XXX/XXXX-XX."""
    c = cnpj.zfill(14)
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:14]}"


def formatar_data_br(data_iso):
    """Converte YYYY-MM-DD para DD/MM/YYYY."""
    if not data_iso or len(data_iso) < 10:
        return data_iso
    parts = data_iso[:10].split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return data_iso


def formatar_valor_br(valor):
    """Formata float para string com vírgula decimal."""
    return f"{valor:.2f}".replace(".", ",")


def contar_registros_csv(csv_path):
    """Conta linhas de dados no CSV (exclui o cabeçalho)."""
    with open(csv_path, encoding="utf-8") as f:
        linhas = f.readlines()
    return sum(1 for l in linhas[1:] if l.strip())


def ler_historico(resumo_path):
    """Lê seção HISTÓRICO DE ATUALIZAÇÕES do resumo existente."""
    if not os.path.isfile(resumo_path):
        return []
    with open(resumo_path, encoding="utf-8") as f:
        content = f.read()
    marker = "HISTÓRICO DE ATUALIZAÇÕES"
    if marker not in content:
        return []
    section = content.split(marker, 1)[1]
    # Pular a linha de separadores (---)
    lines = section.strip().split("\n")
    entries = []
    for line in lines:
        stripped = line.strip()
        if stripped and not stripped.startswith("---"):
            entries.append(stripped)
    return entries


def gerar_resumo(notas, periodo, historico=None):
    """Gera texto de resumo mensal a partir da lista de notas."""
    autorizadas = [n for n in notas if n.get("cstat", "100") == "100"]
    canceladas = [n for n in notas if n.get("cstat") == "110"]
    outros = [n for n in notas if n.get("cstat", "100") not in ("100", "110")]

    total_valor = sum(n["valor_total"] for n in autorizadas)
    com_rateio = [n for n in autorizadas if n["tem_rateio"]]
    sem_rateio = [n for n in autorizadas if not n["tem_rateio"]]
    valor_com_rateio = sum(n["valor_total"] for n in com_rateio)
    valor_sem_rateio = sum(n["valor_total"] for n in sem_rateio)

    valor_por_dia = defaultdict(float)
    notas_por_dia = defaultdict(int)
    for n in autorizadas:
        dia = n["data_servico"][:10] if n["data_servico"] else "sem data"
        valor_por_dia[dia] += n["valor_total"]
        notas_por_dia[dia] += 1

    lines = []
    lines.append(f"RESUMO MENSAL DE NOTAS FISCAIS - {periodo}")
    lines.append("=" * 50)
    lines.append("")
    lines.append("STATUS DAS NOTAS")
    lines.append("-" * 30)
    lines.append(f"  Autorizadas:  {len(autorizadas)}")
    if canceladas:
        lines.append(f"  Canceladas:   {len(canceladas)}")
    if outros:
        lines.append(f"  Outros:       {len(outros)}")
    lines.append(f"  TOTAL:        {len(notas)}")
    lines.append("")
    lines.append("TOTAIS (apenas autorizadas)")
    lines.append("-" * 30)
    lines.append(f"  Total de notas:    {len(autorizadas)}")
    lines.append(f"  Valor total:       R$ {formatar_valor_br(total_valor)}")
    lines.append("")
    lines.append("BREAKDOWN")
    lines.append("-" * 30)
    lines.append(f"  Com rateio:    {len(com_rateio)} notas  |  R$ {formatar_valor_br(valor_com_rateio)}")
    lines.append(f"  Sem rateio:    {len(sem_rateio)} notas  |  R$ {formatar_valor_br(valor_sem_rateio)}")
    lines.append("")
    lines.append("VALOR POR DIA")
    lines.append("-" * 50)
    lines.append(f"  {'Data':<14} {'Notas':>6}   {'Valor':>14}")
    lines.append(f"  {'----':<14} {'-----':>6}   {'---------':>14}")

    for dia in sorted(valor_por_dia.keys()):
        data_fmt = formatar_data_br(dia) if dia != "sem data" else dia
        lines.append(
            f"  {data_fmt:<14} {notas_por_dia[dia]:>6}   R$ {formatar_valor_br(valor_por_dia[dia]):>10}"
        )

    if historico:
        lines.append("")
        lines.append("HISTÓRICO DE ATUALIZAÇÕES")
        lines.append("-" * 50)
        for entry in historico:
            lines.append(f"  {entry}")
        lines.append("")

    lines.append("")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="Extrai notas com cota parte de profissional parceiro")
    parser.add_argument("--output", "-o", default=None,
                        help="Caminho do arquivo CSV de saída")
    parser.add_argument("--periodo", "-p", default=None,
                        help="Filtrar por período (YYYY-MM ou YYYY-MM-DD)")
    parser.add_argument("--max-nsu", type=int, default=None,
                        help="NSU máximo para parar a busca (None = buscar tudo)")
    parser.add_argument("--pdf", nargs="*", default=None,
                        help="PDFs para complementar notas não disponíveis na API")
    parser.add_argument("--senha", "-s", default=None,
                        help="Senha do certificado .pfx (ou use DEPILAGOS_PFX_PASSWORD)")
    args = parser.parse_args()

    # Auto-detectar output pelo período
    if args.output is None:
        if args.periodo:
            parts = args.periodo.split("-")
            if len(parts) >= 2:
                mes_abrev = MESES_ABREV.get(parts[1], parts[1])
                args.output = os.path.join(PROJETO_DIR, "output", f"{mes_abrev}{parts[0]}.csv")
            else:
                args.output = os.path.join(PROJETO_DIR, "output", "notas.csv")
        else:
            args.output = os.path.join(PROJETO_DIR, "output", "notas.csv")

    # Resolver senha: argumento > variável de ambiente > input interativo
    senha = args.senha or os.environ.get("DEPILAGOS_PFX_PASSWORD")
    if not senha:
        import getpass
        senha = getpass.getpass("Senha do certificado .pfx: ")

    print("Extraindo certificado digital...")
    cert_path, key_path = extrair_cert_key(PFX_PATH, senha)

    print("Consultando API NFS-e Nacional...")
    todas_notas = []
    nsu_atual = 1
    total_docs = 0

    while True:
        print(f"  Buscando lote a partir do NSU {nsu_atual}...")
        resposta = consultar_lote(cert_path, key_path, nsu_atual)

        if resposta is None:
            print("  Erro na consulta. Parando.", file=sys.stderr)
            break

        status = resposta.get("StatusProcessamento")
        if status == "NENHUM_DOCUMENTO_LOCALIZADO":
            print("  Nenhum documento encontrado. Fim da busca.")
            break

        if status == "REJEICAO":
            erros = resposta.get("Erros", [])
            for e in erros:
                print(f"  ERRO: {e.get('Descricao', '')} {e.get('Complemento', '')}", file=sys.stderr)
            break

        lote = resposta.get("LoteDFe", [])
        if not lote:
            print("  Lote vazio. Fim da busca.")
            break

        total_docs += len(lote)
        max_nsu_lote = max(d["NSU"] for d in lote)
        print(f"  Recebidos {len(lote)} documentos (NSU {nsu_atual}-{max_nsu_lote})")

        for doc in lote:
            if doc.get("TipoDocumento") != "NFSE":
                continue
            try:
                nota = extrair_dados_nota(doc)
                nota["nsu"] = doc["NSU"]
                todas_notas.append(nota)
            except Exception as e:
                print(f"  Erro processando NSU {doc['NSU']}: {e}", file=sys.stderr)

        nsu_atual = max_nsu_lote + 1

        if args.max_nsu and nsu_atual > args.max_nsu:
            break

    print(f"\nTotal de documentos baixados: {total_docs}")
    print(f"Total de NFS-e processadas: {len(todas_notas)}")

    # Filtrar por período se especificado
    if args.periodo:
        todas_notas = [n for n in todas_notas if n["data_servico"].startswith(args.periodo)]
        print(f"Notas no período {args.periodo}: {len(todas_notas)}")

    # Filtrar apenas notas com rateio
    notas_com_rateio = [n for n in todas_notas if n["tem_rateio"]]
    print(f"Notas com cota parte de profissional (API): {len(notas_com_rateio)}")

    # Complementar com PDFs se fornecidos
    nfse_ids_api = {n["nfse"] for n in todas_notas}

    pdf_paths = args.pdf or []
    # Auto-detectar PDFs no diretório do projeto se --pdf não foi passado
    if not pdf_paths:
        pdf_paths = sorted(glob.glob(os.path.join(PROJETO_DIR, "docs", "Notas*.pdf")))

    if pdf_paths:
        print(f"\nComplementando com {len(pdf_paths)} PDF(s)...")
        notas_pdf = extrair_notas_pdf(pdf_paths)

        # Adicionar apenas notas que NÃO existem na API
        adicionadas = 0
        for nota in notas_pdf:
            if nota["nfse"] not in nfse_ids_api:
                # Aplicar filtro de período se especificado
                if args.periodo and not nota["data_servico"].startswith(args.periodo):
                    continue
                notas_com_rateio.append(nota)
                nfse_ids_api.add(nota["nfse"])
                adicionadas += 1

        if adicionadas:
            print(f"  {adicionadas} notas adicionadas do PDF (não estavam na API)")
        else:
            print("  Todas as notas do PDF já estavam na API")

    print(f"Total notas com cota parte: {len(notas_com_rateio)}")

    # Gerar linhas do CSV
    linhas = []
    for nota in notas_com_rateio:
        is_pdf = nota.get("fonte") == "pdf"
        for cnpj_parceiro, dados in nota["parceiros"].items():
            linhas.append({
                "data_servico": formatar_data_br(nota["data_servico"]),
                "data_emissao": nota["data_emissao"] if is_pdf else formatar_data_br(nota["data_emissao"]),
                "numero_nota": nota["nfse"],
                "cnpj_parceiro": formatar_cnpj(cnpj_parceiro),
                "cota_salao": formatar_valor_br(dados["cota_salao"]),
                "cota_parceiro": formatar_valor_br(dados["cota_parceiro"]),
            })

    linhas.sort(key=lambda x: x["numero_nota"])

    # Determinar path do resumo
    resumo_path = None
    if args.periodo:
        parts = args.periodo.split("-")
        if len(parts) == 2:
            ano, mes = parts
            mes_abrev = MESES_ABREV.get(mes, mes)
            resumo_path = os.path.join(PROJETO_DIR, "output", f"resumo_{mes_abrev}{ano}.txt")
        else:
            resumo_path = args.output.replace(".csv", "_resumo.txt")

    # Comparar com CSV existente e decidir se atualiza
    hoje = datetime.now().strftime("%d/%m/%Y %H:%M")
    historico_existente = ler_historico(resumo_path) if resumo_path else []

    csv_existe = os.path.isfile(args.output)
    registros_novos = len(linhas)

    if csv_existe and args.periodo:
        registros_atual = contar_registros_csv(args.output)
        diff = registros_novos - registros_atual

        if diff == 0:
            print(f"\nCSV já está atualizado ({registros_atual} registros)")
            historico_existente.append(f"{hoje}  Verificação: sem notas novas ({registros_atual} registros)")
        elif diff > 0:
            # Sobrescrever CSV
            _escrever_csv(args.output, linhas)
            print(f"\nCSV atualizado: {registros_atual} → {registros_novos} registros (+{diff} notas)")
            historico_existente.append(
                f"{hoje}  Atualização: {registros_atual} → {registros_novos} registros (+{diff} notas)"
            )
        else:
            # Menos registros - sobrescrever mas avisar
            _escrever_csv(args.output, linhas)
            print(f"\nCSV atualizado: {registros_atual} → {registros_novos} registros ({diff} notas)")
            historico_existente.append(
                f"{hoje}  Atualização: {registros_atual} → {registros_novos} registros ({diff} notas)"
            )
    else:
        _escrever_csv(args.output, linhas)
        print(f"\nCSV salvo em: {args.output}")
        print(f"Total de linhas: {registros_novos}")
        if args.periodo:
            historico_existente.append(f"{hoje}  Primeira extração: {registros_novos} registros")

    # Gerar XLSX compatível com Google Sheets
    xlsx_path = args.output.replace(".csv", ".xlsx")
    _escrever_xlsx(xlsx_path, linhas)

    # Gerar resumo mensal (.txt) - sempre atualizar para registrar histórico
    if args.periodo and resumo_path:
        resumo = gerar_resumo(todas_notas, args.periodo, historico=historico_existente)

        with open(resumo_path, "w", encoding="utf-8") as f:
            f.write(resumo)

        print(f"\nResumo salvo em: {resumo_path}")
        print()
        print(resumo)


def _escrever_csv(path, linhas):
    """Escreve as linhas do CSV no arquivo."""
    with open(path, "w", encoding="utf-8") as f:
        f.write("data servico;data emissao;NUMERO DA NOTA;CNPJ PROFISSIONAL PARCEIRO;"
                "COTA PARTE SALÃO PARCEIRO;COTA PARTE PROFISSIONAL-PARCEIRO\n")
        for r in linhas:
            f.write(f"{r['data_servico']};{r['data_emissao']};{r['numero_nota']};"
                    f"{r['cnpj_parceiro']};{r['cota_salao']};{r['cota_parceiro']}\n")


def _escrever_xlsx(path, linhas):
    """Gera arquivo .xlsx compatível com Google Sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "Notas com Rateio"

    headers = [
        "data servico", "data emissao", "NUMERO DA NOTA",
        "CNPJ PROFISSIONAL PARCEIRO", "COTA PARTE SALÃO PARCEIRO",
        "COTA PARTE PROFISSIONAL-PARCEIRO",
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    for row_idx, r in enumerate(linhas, 2):
        ws.cell(row=row_idx, column=1, value=r["data_servico"])
        ws.cell(row=row_idx, column=2, value=r["data_emissao"])
        ws.cell(row=row_idx, column=3, value=r["numero_nota"])
        ws.cell(row=row_idx, column=4, value=r["cnpj_parceiro"])

        # Cotas como número (float) para Google Sheets poder somar
        cota_salao = float(r["cota_salao"].replace(",", "."))
        cota_parceiro = float(r["cota_parceiro"].replace(",", "."))
        cell_s = ws.cell(row=row_idx, column=5, value=cota_salao)
        cell_p = ws.cell(row=row_idx, column=6, value=cota_parceiro)
        cell_s.number_format = '#.##0,00'
        cell_p.number_format = '#.##0,00'

    # Auto-ajustar largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(path)
    print(f"XLSX salvo em: {path}")


if __name__ == "__main__":
    main()
